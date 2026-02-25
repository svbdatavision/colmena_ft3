#!/usr/bin/env python3
"""Generate a client-ready 3-sheet inventory workbook for CERT AFP."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_modelos_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Modelo": "CERT AFP",
                "Objetivo": (
                    "Predecir probabilidad de aprobación de licencias médicas para priorizar "
                    "auto-aprobación y reducir revisión manual."
                ),
                "Framework Detectado": "LightGBM + scikit-learn + Optuna",
                "Motor SQL/Runtime": "Spark SQL (Databricks)",
                "Tipo (Batch/API)": "Batch",
                "Frecuencia": "Diaria (pipeline); re-proceso on-demand por rango de fechas",
                "Artefacto Modelo Detectado": "fasttrack_model.pkl",
                "Artefacto Features/Transformers": "feature_fasttrack.pkl",
                "Orquestación Detectada": (
                    "run_daily_pipeline.py -> query_lunes/query_diaria + query_2 + FT3_dia.py"
                ),
                "Ubicación Destino (Databricks)": (
                    "/dbfs/FileStore/models/ (configurable por MODELS_PATH/"
                    "FASTTRACK_MODEL_PATH/FEATURE_TRANSFORMERS_PATH)"
                ),
                "Estado Migración": "Migrado a Spark SQL (sin databricks-sql-connector)",
                "Observaciones": (
                    "Artefactos .pkl no presentes en /workspace/models al momento del relevamiento."
                ),
            }
        ]
    )


def build_inputs_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Modelo": "CERT AFP",
                "Tipo Input": "Tabla SQL (ingesta diaria)",
                "Nombre": "SBN_LM_INPUT_DIARIO_ALFIL",
                "Ruta Origen Detectada": "query_diaria.sql / query_lunes.sql",
                "Tabla SQL": "OPX.P_DDV_OPX_MDPREDICTIVO.SBN_LM_INPUT_DIARIO_ALFIL",
                "Falta": "No",
                "Ubicación Destino en Databricks": (
                    "OPX.P_DDV_OPX_MDPREDICTIVO.SBN_LM_INPUT_DIARIO_ALFIL"
                ),
                "Observaciones": "Fuente principal de licencias nuevas.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Input": "Tabla SQL (histórico base)",
                "Nombre": "CPA_LM_BASE_AMPLIADA",
                "Ruta Origen Detectada": "query_diaria.sql / query_lunes.sql",
                "Tabla SQL": "OPX.P_DDV_OPX_MDPREDICTIVO.CPA_LM_BASE_AMPLIADA",
                "Falta": "No",
                "Ubicación Destino en Databricks": "OPX.P_DDV_OPX_MDPREDICTIVO.CPA_LM_BASE_AMPLIADA",
                "Observaciones": "Se usa para consolidación/atributos históricos.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Input": "Tabla SQL (feature table intermedia)",
                "Nombre": "MODELO_LM_202507_OPTIMIZADO",
                "Ruta Origen Detectada": "query_diaria.sql / query_lunes.sql / query_2.sql",
                "Tabla SQL": "OPX.P_DDV_OPX_MDPREDICTIVO.MODELO_LM_202507_OPTIMIZADO",
                "Falta": "No",
                "Ubicación Destino en Databricks": (
                    "OPX.P_DDV_OPX_MDPREDICTIVO.MODELO_LM_202507_OPTIMIZADO"
                ),
                "Observaciones": "Input directo para construir tabla TRAIN.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Input": "Tabla SQL (training/scoring)",
                "Nombre": "MODELO_LM_202507_TRAIN",
                "Ruta Origen Detectada": "query_2.sql / FT3_dia.py / FT30.py",
                "Tabla SQL": "OPX.P_DDV_OPX_MDPREDICTIVO.MODELO_LM_202507_TRAIN",
                "Falta": "No",
                "Ubicación Destino en Databricks": "OPX.P_DDV_OPX_MDPREDICTIVO.MODELO_LM_202507_TRAIN",
                "Observaciones": "Tabla principal consumida por scoring diario.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Input": "Archivo artefacto ML",
                "Nombre": "fasttrack_model.pkl",
                "Ruta Origen Detectada": (
                    "models/fasttrack_model.pkl (y fallback en /dbfs/... por _resolve_artifact_path)"
                ),
                "Tabla SQL": "-",
                "Falta": "Sí (no detectado en /workspace/models)",
                "Ubicación Destino en Databricks": "/dbfs/FileStore/models/fasttrack_model.pkl",
                "Observaciones": "Requerido por FT3_dia.py y FT30.py para inferencia.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Input": "Archivo transformers ML",
                "Nombre": "feature_fasttrack.pkl",
                "Ruta Origen Detectada": (
                    "models/feature_fasttrack.pkl (y fallback en /dbfs/... por _resolve_artifact_path)"
                ),
                "Tabla SQL": "-",
                "Falta": "Sí (no detectado en /workspace/models)",
                "Ubicación Destino en Databricks": "/dbfs/FileStore/models/feature_fasttrack.pkl",
                "Observaciones": "Requerido para aplicar encoding/TF-IDF en scoring.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Input": "Archivo configuración",
                "Nombre": "config.yaml",
                "Ruta Origen Detectada": "/workspace/config.yaml",
                "Tabla SQL": "data.snowflake.database/schema + target/model params",
                "Falta": "No",
                "Ubicación Destino en Databricks": "/Workspace/Repos/.../config.yaml (o ruta de job)",
                "Observaciones": "Define DB/schema y parámetros del modelo.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Input": "Archivo definición de features",
                "Nombre": "Variables_cat_train.py",
                "Ruta Origen Detectada": "/workspace/Variables_cat_train.py",
                "Tabla SQL": "-",
                "Falta": "No",
                "Ubicación Destino en Databricks": "/Workspace/Repos/.../Variables_cat_train.py",
                "Observaciones": (
                    "Fuente de grupos de variables usados por feature_engineering.py."
                ),
            },
        ]
    )


def build_outputs_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Modelo": "CERT AFP",
                "Tipo Output": "Tabla SQL (dataset train actualizado)",
                "Nombre": "MODELO_LM_202507_TRAIN",
                "Ubicación Actual Detectada": (
                    "OPX.P_DDV_OPX_MDPREDICTIVO.MODELO_LM_202507_TRAIN (query_2.sql)"
                ),
                "Ubicación Destino en Databricks": (
                    "OPX.P_DDV_OPX_MDPREDICTIVO.MODELO_LM_202507_TRAIN"
                ),
                "Consumidor": "FT3_dia.py / FT30.py",
                "Observaciones": "Se recrea en cada corrida de query_2.sql.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Output": "Tabla SQL (predicciones diarias)",
                "Nombre": "FT30_PREDICCIONES_DIARIAS",
                "Ubicación Actual Detectada": (
                    "Creada/escrita por FT3_dia.py y FT30.py (calificada con database.schema)"
                ),
                "Ubicación Destino en Databricks": (
                    "OPX.P_DDV_OPX_MDPREDICTIVO.FT30_PREDICCIONES_DIARIAS"
                ),
                "Consumidor": "Operación/seguimiento",
                "Observaciones": "Incluye con dictamen + pendientes; MERGE para evitar duplicados.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Output": "Tabla SQL (pendientes)",
                "Nombre": "FT30_LICENCIAS_PENDIENTES",
                "Ubicación Actual Detectada": (
                    "Creada/escrita por FT3_dia.py y FT30.py (calificada con database.schema)"
                ),
                "Ubicación Destino en Databricks": (
                    "OPX.P_DDV_OPX_MDPREDICTIVO.FT30_LICENCIAS_PENDIENTES"
                ),
                "Consumidor": "Operación de priorización",
                "Observaciones": "Solo licencias pendientes; actualizada por MERGE.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Output": "Archivo reporte diario",
                "Nombre": "reporte_dia_YYYYMMDD_HHMMSS.(xlsx|csv)",
                "Ubicación Actual Detectada": "results/ (creado automáticamente desde FT3_dia.py)",
                "Ubicación Destino en Databricks": "/Workspace/Repos/.../results o DBFS según job",
                "Consumidor": "Negocio / monitoreo",
                "Observaciones": "Si falta openpyxl, genera CSV de fallback.",
            },
            {
                "Modelo": "CERT AFP",
                "Tipo Output": "Archivo reporte rango",
                "Nombre": "reporte_completo_YYYYMMDD_HHMMSS.(xlsx|csv)",
                "Ubicación Actual Detectada": "results/ (creado automáticamente desde FT30.py)",
                "Ubicación Destino en Databricks": "/Workspace/Repos/.../results o DBFS según job",
                "Consumidor": "Análisis ad-hoc",
                "Observaciones": (
                    "Genera adicionales: *_licencias_pendientes.csv y *_con_dictamen.csv."
                ),
            },
        ]
    )


def format_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_cells in ws.columns:
            col_letter = col_cells[0].column_letter
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
            ws.column_dimensions[col_letter].width = min(max(14, max_len + 2), 70)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)


def main() -> int:
    output_dir = Path("/workspace/deliverables")
    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / "CERT_AFP_modelo_inputs_outputs_databricks.xlsx"

    modelos_df = build_modelos_df()
    inputs_df = build_inputs_df()
    outputs_df = build_outputs_df()

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        modelos_df.to_excel(writer, sheet_name="Modelos", index=False)
        inputs_df.to_excel(writer, sheet_name="Inputs", index=False)
        outputs_df.to_excel(writer, sheet_name="Outputs", index=False)

    format_workbook(out_path)
    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
